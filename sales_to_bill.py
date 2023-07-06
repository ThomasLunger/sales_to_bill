import pandas as pd
import os
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import PatternFill

# get file paths
sales_path = input("Enter path to Sales Sheet: ")
bill_path = input("Enter path to Bill Trigger: ")

# read excel sheets
sales_sheet = pd.read_excel(sales_path)
bill_sheet = pd.read_excel(bill_path)

# convert 'Ship Date' column to string format
sales_sheet['Ship Date'] = sales_sheet['Ship Date'].astype(str)

# define function to add 5 days to a date
def add_5_days(date_str):
    formats = ['%m/%d/%Y', '%m/%d/%y', '%Y-%m-%d %H:%M:%S']
    for fmt in formats:
        try:
            date = datetime.strptime(date_str, fmt)
            break
        except ValueError:
            pass
    else:
        raise ValueError(f"Could not parse date: {date_str}")
    new_date = (date + timedelta(days=5)).strftime('%m/%d/%Y')
    return new_date

# define function to color rows green
def color_rows(row):
    bg_color = '#d3d3d3' if row['In Service Date'] == '' else '#00ff00'
    return ['background-color: %s' % bg_color] * len(row)

# loop through each row in sales sheet and update in service date in bill trigger
updated_rows = []
for index, row in sales_sheet.iterrows():
    po_number = row['PO #']
    ship_dates = row['Ship Date']
    box_serial_numbers = row['Box Serial Number(s)']
    ship_tracking_numbers = row['Ship Tracking#']
    matching_rows = bill_sheet[bill_sheet['AT&T PO #'] == po_number]
    if len(matching_rows) > 0:
        print(f"Matching PO#: {po_number}")
        ship_dates_list = ship_dates.split('\n')
        new_dates = []
        for date in ship_dates_list:
            try:
                new_date = add_5_days(date)
                new_dates.append(new_date)
                print(f"Ship date: {date} -> New In Service Date: {new_date}")
            except ValueError:
                print(f"Could not parse date: {date}")
        # apply updates only if new_dates list is not empty
        if new_dates:
            bill_sheet.loc[matching_rows.index, 'In Service Date'] = '\n'.join(new_dates)
            bill_sheet.loc[matching_rows.index, 'Serial #'] = box_serial_numbers
            bill_sheet.loc[matching_rows.index, 'Tracking #'] = ship_tracking_numbers
            updated_rows.extend(matching_rows.index.tolist())

# apply styling to updated bill trigger
styled_bill_sheet = bill_sheet.style.apply(color_rows, axis=1)

# get current date and time to append to filename
now = datetime.now()
dt_string = now.strftime("%Y-%m-%d_%H-%M-%S")

# save updated bill trigger
filename, ext = os.path.splitext(bill_path)
new_filename = f"{filename}_{dt_string}{ext}"
styled_bill_sheet.to_excel(new_filename, index=False)
print(f"Updated Bill Trigger saved as {new_filename}")

# load the workbook
workbook = openpyxl.load_workbook(new_filename)

# select the worksheet
worksheet = workbook.active

# define fill colors
green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
light_grey_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

# loop through the rows in the worksheet
for row in worksheet.iter_rows(min_row=2):
    in_service_date = row[10].value
    fill_color = green_fill if row[0].fill.start_color.rgb == '00FF00' else None
    if in_service_date in ('', None):
        fill_color = light_grey_fill
    for cell in row:
        if fill_color:
            cell.fill = fill_color
            
# save the updated workbook
workbook.save(new_filename)
# print a message to confirm completion
print("Updated Bill Trigger saved with colored rows.")
os.startfile(new_filename)
